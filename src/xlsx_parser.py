"""
XLSX file parser for Sprinklr message exports.

Parses pre-exported XLSX files and builds an indexed message store
for efficient lookups during hybrid ingestion.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Iterator
import openpyxl


# Column name to index mapping for Sprinklr XLSX exports
# Column indices are 1-based (openpyxl convention)
# Discovered from actual XLSX headers using discover_xlsx_columns()
COLUMN_MAP = {
    "message": 6,               # Message content
    "social_network": 1,        # SocialNetwork (e.g., INSTAGRAM, FACEBOOK)
    "created_time": 7,          # CreatedTime (timestamp string)
    "sender_screen_name": 3,    # SenderScreenName
    "receiver_screen_name": 10, # ReceiverScreenName (often the brand account)
    "message_type": 117,        # Message Type (e.g., "Facebook Messenger ( Sent )")
    "conversation_id": 107,     # ConversationId
    "profile_sprinklr_id": 81,  # Profile Sprinklr ID
    "brand": 40,                # Brand (Activity, Product)
    "sentiment": 16,            # Sentiment (POSITIVE/NEGATIVE/NEUTRAL)
    "language": 23,             # YHI Localisation Language
    "associated_cases": 109,    # Associated Cases (case number)
    "country": 111,             # Country
}


class XlsxParser:
    """
    Parser for Sprinklr XLSX message exports.

    Builds an in-memory index of messages keyed by case number
    for fast lookup during hybrid ingestion.
    """

    def __init__(self, xlsx_directory: Optional[str] = None):
        """
        Initialize the parser and optionally load files.

        Args:
            xlsx_directory: Directory containing XLSX files to load.
                           If None, must call load_directory() later.
        """
        # Message index: case_number -> list of messages
        self._messages_by_case: Dict[int, List[Dict[str, Any]]] = {}

        # Statistics
        self._total_messages = 0
        self._total_files = 0
        self._load_errors: List[str] = []

        if xlsx_directory:
            self.load_directory(xlsx_directory)

    def load_directory(self, xlsx_directory: str) -> int:
        """
        Load all XLSX files from a directory.

        Args:
            xlsx_directory: Path to directory containing XLSX files

        Returns:
            Number of messages loaded
        """
        directory = Path(xlsx_directory)
        if not directory.exists():
            raise FileNotFoundError(f"Directory not found: {xlsx_directory}")

        # Find all XLSX files recursively
        xlsx_files = list(directory.rglob("*.xlsx"))

        if not xlsx_files:
            print(f"Warning: No XLSX files found in {xlsx_directory}")
            return 0

        print(f"Found {len(xlsx_files)} XLSX files to process...")

        for i, xlsx_file in enumerate(xlsx_files, 1):
            print(f"  [{i}/{len(xlsx_files)}] Loading {xlsx_file.name}...")
            try:
                self._load_file(xlsx_file)
                self._total_files += 1
            except Exception as e:
                error_msg = f"Error loading {xlsx_file.name}: {e}"
                self._load_errors.append(error_msg)
                print(f"    Warning: {error_msg}")

        print(f"\nLoaded {self._total_messages:,} messages from {self._total_files} files")
        print(f"Indexed {len(self._messages_by_case):,} unique cases")

        return self._total_messages

    def _load_file(self, xlsx_path: Path) -> int:
        """
        Load messages from a single XLSX file.

        Args:
            xlsx_path: Path to XLSX file

        Returns:
            Number of messages loaded from this file
        """
        # Load workbook in read-only mode for efficiency
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active

        messages_loaded = 0

        # Skip header row
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                message = self._parse_row(row)
                if message and message.get("case_number"):
                    case_number = message["case_number"]

                    if case_number not in self._messages_by_case:
                        self._messages_by_case[case_number] = []

                    self._messages_by_case[case_number].append(message)
                    messages_loaded += 1
                    self._total_messages += 1

            except Exception as e:
                # Skip problematic rows but continue processing
                pass

        wb.close()
        return messages_loaded

    def _parse_row(self, row: tuple) -> Optional[Dict[str, Any]]:
        """
        Parse a single row into a message dictionary.

        Args:
            row: Tuple of cell values from the row

        Returns:
            Message dictionary or None if row is empty/invalid
        """
        if not row or len(row) < COLUMN_MAP["associated_cases"]:
            return None

        # Get case number (column index is 1-based, tuple is 0-based)
        case_number_raw = row[COLUMN_MAP["associated_cases"] - 1]
        if not case_number_raw:
            return None

        # Parse case number (handle "123" or 123)
        try:
            case_number = int(str(case_number_raw).strip())
        except (ValueError, TypeError):
            return None

        # Get message content
        message_content = row[COLUMN_MAP["message"] - 1] if len(row) >= COLUMN_MAP["message"] else None
        if not message_content:
            message_content = ""
        else:
            message_content = str(message_content).strip()

        # Parse created time
        created_time_raw = row[COLUMN_MAP["created_time"] - 1] if len(row) >= COLUMN_MAP["created_time"] else None
        created_time_epoch = self._parse_datetime(created_time_raw)

        # Get message type for role detection
        message_type_raw = row[COLUMN_MAP["message_type"] - 1] if len(row) >= COLUMN_MAP["message_type"] else ""
        message_type = str(message_type_raw).strip() if message_type_raw else ""

        # Determine role from Message Type column
        # "Sent" in message type indicates outbound/brand message
        # Everything else is inbound/user message
        is_sent = "sent" in message_type.lower() or "outbound" in message_type.lower()
        role = "agent" if is_sent else "user"

        # Get sender name
        sender = row[COLUMN_MAP["sender_screen_name"] - 1] if len(row) >= COLUMN_MAP["sender_screen_name"] else None
        if sender:
            sender = str(sender).strip()
        else:
            sender = "Unknown"

        # Get other metadata with safe indexing
        def safe_get(col_name):
            idx = COLUMN_MAP.get(col_name, 0)
            if idx > 0 and len(row) >= idx:
                return row[idx - 1]
            return None

        social_network = safe_get("social_network") or ""
        brand = safe_get("brand") or ""
        sentiment = safe_get("sentiment") or ""
        language = safe_get("language") or ""
        conversation_id = safe_get("conversation_id") or ""
        profile_id = safe_get("profile_sprinklr_id") or ""
        country = safe_get("country") or ""

        return {
            "case_number": case_number,
            "content": message_content,
            "role": role,
            "sender": sender,
            "created_time_epoch": created_time_epoch,
            "social_network": str(social_network).strip(),
            "brand": str(brand).strip(),
            "sentiment": str(sentiment).strip(),
            "language": str(language).strip(),
            "message_type": message_type,
            "conversation_id": str(conversation_id).strip() if conversation_id else "",
            "profile_id": str(profile_id).strip() if profile_id else "",
            "country": str(country).strip() if country else "",
        }

    def _parse_datetime(self, dt_value: Any) -> Optional[int]:
        """
        Parse datetime value to epoch milliseconds.

        Handles various formats:
        - datetime object
        - String like "2025-01-15 10:30:00"
        - Epoch timestamp (already numeric)

        Args:
            dt_value: Raw datetime value from cell

        Returns:
            Epoch milliseconds or None
        """
        if dt_value is None:
            return None

        # Already a datetime
        if isinstance(dt_value, datetime):
            return int(dt_value.timestamp() * 1000)

        # Numeric (could be epoch)
        if isinstance(dt_value, (int, float)):
            # If it looks like epoch ms (> 1 trillion), use as-is
            if dt_value > 1_000_000_000_000:
                return int(dt_value)
            # If it looks like epoch seconds
            elif dt_value > 1_000_000_000:
                return int(dt_value * 1000)
            return None

        # String - try to parse
        if isinstance(dt_value, str):
            dt_str = dt_value.strip()
            if not dt_str:
                return None

            # Try common formats
            formats = [
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%dT%H:%M:%S.%f",
                "%m/%d/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M:%S",
            ]

            for fmt in formats:
                try:
                    dt = datetime.strptime(dt_str, fmt)
                    return int(dt.timestamp() * 1000)
                except ValueError:
                    continue

        return None

    def get_messages_for_case(self, case_number: int) -> List[Dict[str, Any]]:
        """
        Get all messages for a case number, sorted by timestamp.

        Args:
            case_number: The case number to look up

        Returns:
            List of message dictionaries sorted by created_time
        """
        messages = self._messages_by_case.get(case_number, [])

        # Sort by created_time_epoch
        return sorted(
            messages,
            key=lambda m: m.get("created_time_epoch") or 0
        )

    def has_case(self, case_number: int) -> bool:
        """
        Check if messages exist for a case.

        Args:
            case_number: The case number to check

        Returns:
            True if messages exist for this case
        """
        return case_number in self._messages_by_case

    def get_case_numbers(self) -> List[int]:
        """
        Get all case numbers with messages.

        Returns:
            List of case numbers
        """
        return list(self._messages_by_case.keys())

    @property
    def message_count(self) -> int:
        """Total number of messages loaded."""
        return self._total_messages

    @property
    def case_count(self) -> int:
        """Total number of unique cases."""
        return len(self._messages_by_case)

    @property
    def file_count(self) -> int:
        """Number of XLSX files loaded."""
        return self._total_files

    @property
    def load_errors(self) -> List[str]:
        """List of errors encountered during loading."""
        return self._load_errors.copy()

    def get_stats(self) -> Dict[str, Any]:
        """
        Get statistics about the loaded data.

        Returns:
            Dictionary with statistics
        """
        # Calculate messages per case distribution
        messages_per_case = [len(msgs) for msgs in self._messages_by_case.values()]

        return {
            "total_messages": self._total_messages,
            "total_cases": len(self._messages_by_case),
            "files_loaded": self._total_files,
            "errors": len(self._load_errors),
            "avg_messages_per_case": sum(messages_per_case) / len(messages_per_case) if messages_per_case else 0,
            "max_messages_per_case": max(messages_per_case) if messages_per_case else 0,
            "min_messages_per_case": min(messages_per_case) if messages_per_case else 0,
        }


def discover_xlsx_columns(xlsx_path: str) -> Dict[str, int]:
    """
    Utility function to discover column names and indices in an XLSX file.

    Useful for debugging and mapping new XLSX export formats.

    Args:
        xlsx_path: Path to XLSX file

    Returns:
        Dictionary mapping column names to 1-based indices
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    columns = {}
    # Use iter_rows with values_only to get actual cell values
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for col_num, value in enumerate(row, start=1):
            if value:
                columns[str(value).strip()] = col_num

    wb.close()
    return columns
